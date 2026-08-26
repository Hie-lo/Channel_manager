"""
ساخت فایل‌های نمونه اکسل برای همه کسب‌وکارها (به‌جز computer_shop)

اجرا:
    python scripts/generate_excel_templates.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "app" / "business" / "templates"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ─── رنگ‌های استاندارد ───────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # آبی تیره
CLR_HEADER_FONT = "FFFFFF"   # سفید
CLR_REQ_BG      = "FFF2CC"   # زرد کم‌رنگ (ستون اجباری)
CLR_OPT_BG      = "F2F2F2"   # خاکستری روشن (ستون اختیاری)
CLR_GUIDE_BG    = "E2EFDA"   # سبز بسیار کم‌رنگ (شیت راهنما)
CLR_DATA_ROW_1  = "FFFFFF"
CLR_DATA_ROW_2  = "EBF3FB"   # آبی خیلی کم‌رنگ (ردیف‌های در هم)


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row: int, col: int, value: str, required: bool = True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=CLR_HEADER_FONT, size=11)
    cell.fill      = PatternFill("solid", fgColor=CLR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def _req_label(ws, row: int, col: int, value: str):
    """برچسب زیر هدر برای ستون اجباری"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(italic=True, color="C00000", size=9)
    cell.fill      = PatternFill("solid", fgColor=CLR_REQ_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()


def _opt_label(ws, row: int, col: int, value: str):
    """برچسب زیر هدر برای ستون اختیاری"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(italic=True, color="595959", size=9)
    cell.fill      = PatternFill("solid", fgColor=CLR_OPT_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()


def _data_cell(ws, row: int, col: int, value, even_row: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    bg   = CLR_DATA_ROW_2 if even_row else CLR_DATA_ROW_1
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border    = _thin_border()
    return cell


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _freeze_header(ws, freeze_below_row: int = 3):
    ws.freeze_panes = ws.cell(row=freeze_below_row, column=1)


def _add_guide_sheet(wb: Workbook, business_name: str, sheets_info: list[dict]):
    """شیت راهنما"""
    ws = wb.create_sheet("راهنما")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "70AD47"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50

    def hdr(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=True, color="FFFFFF", size=12)
        cell.fill      = PatternFill("solid", fgColor="375623")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

    def row(r, c, v, bold=False):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill      = PatternFill("solid", fgColor=CLR_GUIDE_BG)
        cell.font      = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
        cell.protection = Protection(locked=False)

    ws.row_dimensions[1].height = 28
    hdr(1, 1, "بخش")
    hdr(1, 2, "توضیحات")

    items = [
        ("کسب‌وکار", business_name, True),
        ("ساختار فایل", f"این فایل دارای {len(sheets_info)} شیت داده است.", False),
    ]
    for info in sheets_info:
        items.append((f"شیت: {info['name']}", info['desc'], False))

    items += [
        ("قیمت", "عدد صحیح به تومان — بدون کاما یا متن", False),
        ("موجودی", "عدد صحیح مثبت — ۰ به‌معنای ناموجود است", False),
        ("کد محصول", "باید منحصر‌به‌فرد باشد — تکراری موجب رونویسی می‌شود", False),
        ("لینک عکس", "آدرس URL کامل عکس محصول (اختیاری)", False),
        ("ردیف اول", "ردیف اول هر شیت هدر است — تغییر ندهید", True),
    ]

    for i, (k, v, b) in enumerate(items, start=2):
        ws.row_dimensions[i].height = 22
        row(i, 1, k, bold=b)
        row(i, 2, v, bold=False)

    ws.sheet_state = "visible"


# ═══════════════════════════════════════════════════════════════════════════════
# mobile_shop.xlsx
# شیت: smartphones
# فیلدهای اجباری: کد محصول، نام محصول، برند، حافظه رم، حافظه داخلی، قیمت، موجودی
# فیلدهای اختیاری: دوربین، باتری، رنگ، توضیحات، لینک عکس
# ═══════════════════════════════════════════════════════════════════════════════

def build_mobile_shop():
    wb = Workbook()
    wb.remove(wb.active)          # حذف Sheet پیش‌فرض

    ws = wb.create_sheet("smartphones")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "4472C4"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    headers = [
        ("کد محصول",    True,  16),
        ("نام محصول",   True,  30),
        ("برند",         True,  14),
        ("حافظه رم",    True,  14),
        ("حافظه داخلی", True,  16),
        ("قیمت",         True,  14),
        ("موجودی",       True,  12),
        ("دوربین",       False, 18),
        ("باتری",        False, 14),
        ("رنگ",          False, 14),
        ("توضیحات",      False, 30),
        ("لینک عکس",    False, 35),
    ]

    for col, (name, required, width) in enumerate(headers, start=1):
        _header_cell(ws, 1, col, name, required)
        label = "* اجباری" if required else "اختیاری"
        if required:
            _req_label(ws, 2, col, label)
        else:
            _opt_label(ws, 2, col, label)
        _set_col_width(ws, col, width)

    # ردیف‌های نمونه (۸ محصول)
    sample_data = [
        ("MOB-001", "سامسونگ Galaxy A55",  "Samsung", "8GB",  "256GB", 18500000, 12, "64+12+5 مگاپیکسل", "5000 mAh", "آبی نیلی",      "گوشی میان‌رده با صفحه‌نمایش Super AMOLED",   ""),
        ("MOB-002", "سامسونگ Galaxy S24",  "Samsung", "8GB",  "256GB", 39900000,  5, "50+10+12 مگاپیکسل","4000 mAh", "بنفش",          "پرچم‌دار ۲۰۲۴ با پردازنده Snapdragon 8 Gen 3", ""),
        ("MOB-003", "شیائومی Redmi Note 13","Xiaomi",  "8GB",  "128GB",  9800000, 20, "108+8+2 مگاپیکسل", "5000 mAh", "مشکی",          "مناسب برای کاربران روزمره",                   ""),
        ("MOB-004", "اپل iPhone 15",        "Apple",   "6GB",  "128GB", 54000000,  3, "48+12 مگاپیکسل",   "3877 mAh", "طلایی",         "آیفون ۱۵ با تراشه A16 Bionic",               ""),
        ("MOB-005", "اپل iPhone 15 Pro",    "Apple",   "8GB",  "256GB", 72000000,  2, "48+12+12 مگاپیکسل","3274 mAh", "تیتانیوم طبیعی","آیفون پرو با فریم تیتانیوم و دکمه Action",   ""),
        ("MOB-006", "وانپلاس 12",           "OnePlus", "12GB", "256GB", 38000000,  7, "50+48+64 مگاپیکسل","5400 mAh", "سیلور",         "شارژ سریع ۱۰۰ واتی SUPERVOOC",               ""),
        ("MOB-007", "شیائومی 14T Pro",      "Xiaomi",  "12GB", "512GB", 32000000,  4, "50+50+12 مگاپیکسل","5000 mAh", "مشکی تیتانیوم", "همکاری با Leica برای دوربین",                 ""),
        ("MOB-008", "سامسونگ Galaxy Z Fold6","Samsung","12GB","512GB", 95000000,  1, "50+10+12 مگاپیکسل","4400 mAh", "نقره‌ای",        "گوشی تاشو نسل ششم",                          ""),
    ]

    for i, row_data in enumerate(sample_data, start=3):
        even = (i % 2 == 0)
        for col, val in enumerate(row_data, start=1):
            _data_cell(ws, i, col, val, even)

    _freeze_header(ws, freeze_below_row=3)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _add_guide_sheet(wb, "فروشگاه موبایل و تبلت", [
        {"name": "smartphones", "desc": "لیست گوشی‌های موبایل — هر ردیف یک مدل"},
    ])

    path = OUTPUT_DIR / "mobile_shop.xlsx"
    wb.save(path)
    print(f"  ✅  {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# clothing_shop.xlsx
# شیت: clothing
# فیلدهای اجباری: کد محصول، نام محصول، برند، سایزبندی، قیمت، موجودی
# فیلدهای اختیاری: رنگ‌بندی، جنس، توضیحات، لینک عکس
# ═══════════════════════════════════════════════════════════════════════════════

def build_clothing_shop():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("clothing")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "E36C09"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    headers = [
        ("کد محصول",  True,  16),
        ("نام محصول", True,  30),
        ("برند",       True,  14),
        ("سایزبندی",  True,  16),
        ("قیمت",       True,  14),
        ("موجودی",     True,  12),
        ("رنگ‌بندی",  False, 20),
        ("جنس",        False, 16),
        ("توضیحات",    False, 35),
        ("لینک عکس",  False, 35),
    ]

    for col, (name, required, width) in enumerate(headers, start=1):
        _header_cell(ws, 1, col, name, required)
        label = "* اجباری" if required else "اختیاری"
        if required:
            _req_label(ws, 2, col, label)
        else:
            _opt_label(ws, 2, col, label)
        _set_col_width(ws, col, width)

    sample_data = [
        ("CL-001", "پیراهن مردانه آستین بلند",  "Zara",     "S,M,L,XL",    450000, 12, "سفید، مشکی، آبی",   "پنبه ۱۰۰٪",       "مناسب برای محیط کار و دورهمی",            ""),
        ("CL-002", "شلوار جین مردانه",           "Levi's",   "30,32,34,36", 680000,  8, "آبی روشن، آبی تیره","دنیم",             "جین کلاسیک پنج‌جیبه",                      ""),
        ("CL-003", "بلوز زنانه یقه گرد",         "Mango",    "XS,S,M",      320000,  5, "صورتی، سفید",        "پنبه ویسکوز",      "طرح ساده مناسب فصل بهار",                  ""),
        ("CL-004", "هودی مردانه زیپ‌دار",        "Nike",     "M,L,XL,XXL",  890000,  6, "خاکستری، مشکی",     "فلیس",             "گرم و سبک با جیب کانگورو",                 ""),
        ("CL-005", "دامن میدی",                  "Zara",     "XS,S,M,L",    520000,  9, "قرمز، بژ، سبز",      "ساتن",             "دامن نیم‌بلند با چین‌های ظریف",             ""),
        ("CL-006", "تی‌شرت ورزشی",               "Adidas",   "S,M,L,XL",    280000, 20, "سفید، مشکی، آبی",   "پلی‌استر",         "مناسب باشگاه و پیاده‌روی",                  ""),
        ("CL-007", "کاپشن زنانه پَفی",           "LC Waikiki","M,L,XL",    1200000,  3, "سرمه‌ای، کرم",       "پلی‌استر پَد",     "گرم و سبک برای فصل زمستان",               ""),
        ("CL-008", "شلوارک ورزشی",               "Bershka",  "S,M,L",       190000, 15, "مشکی، خاکستری",     "پنبه پلی‌استر",    "مناسب خانه و ورزش",                        ""),
        ("CL-009", "کت اسپرت",                   "Mango",    "36,38,40,42", 750000,  7, "کرم، خاکستری",       "پشم مصنوعی",       "کت رسمی با دکمه‌های کلاسیک",               ""),
        ("CL-010", "پیراهن بلند زنانه",          "H&M",      "XS,S,M,L,XL", 520000,  4, "گلبهی، سفید",        "کتان",             "پیراهن بلند مناسب تابستان",                ""),
    ]

    for i, row_data in enumerate(sample_data, start=3):
        even = (i % 2 == 0)
        for col, val in enumerate(row_data, start=1):
            _data_cell(ws, i, col, val, even)

    _freeze_header(ws, freeze_below_row=3)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _add_guide_sheet(wb, "فروشگاه پوشاک و کفش", [
        {"name": "clothing", "desc": "لیست پوشاک — هر ردیف یک محصول لباس یا کفش"},
    ])

    path = OUTPUT_DIR / "clothing_shop.xlsx"
    wb.save(path)
    print(f"  ✅  {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# other.xlsx
# شیت: products
# فیلدهای اجباری: کد محصول، نام محصول، قیمت، موجودی
# فیلدهای اختیاری: برند، توضیحات، لینک عکس
# ═══════════════════════════════════════════════════════════════════════════════

def build_other():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("products")
    ws.sheet_view.rightToLeft = True
    ws.sheet_properties.tabColor = "7030A0"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    headers = [
        ("کد محصول",  True,  16),
        ("نام محصول", True,  35),
        ("قیمت",       True,  14),
        ("موجودی",     True,  12),
        ("برند",       False, 16),
        ("توضیحات",    False, 40),
        ("لینک عکس",  False, 35),
    ]

    for col, (name, required, width) in enumerate(headers, start=1):
        _header_cell(ws, 1, col, name, required)
        label = "* اجباری" if required else "اختیاری"
        if required:
            _req_label(ws, 2, col, label)
        else:
            _opt_label(ws, 2, col, label)
        _set_col_width(ws, col, width)

    sample_data = [
        ("PRD-001", "رز هلندی دسته ۲۰ شاخه",  250000, 30, "گل‌فروشی بهار", "تازه، مناسب هدیه",                         ""),
        ("PRD-002", "کتاب صد سال تنهایی",       180000, 15, "انتشارات نیلوفر","ترجمه فارسی، جلد سخت",                    ""),
        ("PRD-003", "عسل طبیعی یک کیلویی",      420000, 25, "کندوی سبز",      "عسل آویشن کوهی، بدون افزودنی",            ""),
        ("PRD-004", "جاشمعی سرامیکی دستساز",   95000,  40, "کارگاه سفال",    "رنگ‌بندی متنوع، ارتفاع ۸ سانتی‌متر",     ""),
        ("PRD-005", "چای ایرانی ۵۰۰ گرمی",     145000, 60, "باغ چای لاهیجان","چای سر‌گل درجه یک لاهیجان",               ""),
        ("PRD-006", "تابلو نقاشی آبرنگ",        380000,  8, "گالری هنر نو",   "اثر هنری اصیل، ۳۰×۴۰ سانتی‌متر",         ""),
        ("PRD-007", "ماگ سرامیکی طرح‌دار",       75000,  50, "دیجی‌گیفت",     "ظرفیت ۳۵۰ میلی‌لیتر، ماشین‌شور",         ""),
        ("PRD-008", "صابون طبیعی لوندر",         55000,  80, "طبیعت‌گرا",      "صابون سرد، بدون مواد مصنوعی",             ""),
        ("PRD-009", "ست قلم فلزی",              220000, 12, "پن استار",       "ست ۵ عددی با جعبه هدیه",                   ""),
        ("PRD-010", "پادری پارچه‌ای",           165000, 35, "خانه‌آرا",        "ضد لغزش، ابعاد ۶۰×۴۰ سانتی‌متر",         ""),
    ]

    for i, row_data in enumerate(sample_data, start=3):
        even = (i % 2 == 0)
        for col, val in enumerate(row_data, start=1):
            _data_cell(ws, i, col, val, even)

    _freeze_header(ws, freeze_below_row=3)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    _add_guide_sheet(wb, "فروشگاه عمومی (سایر)", [
        {"name": "products", "desc": "لیست محصولات — برای هر نوع فروشگاهی قابل استفاده است"},
    ])

    path = OUTPUT_DIR / "other.xlsx"
    wb.save(path)
    print(f"  ✅  {path}")


# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("⏳ در حال ساخت فایل‌های نمونه اکسل...")
    build_mobile_shop()
    build_clothing_shop()
    build_other()
    print("✅ همه فایل‌ها ساخته شدند.")
