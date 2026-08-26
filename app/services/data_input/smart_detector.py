"""
موتور تشخیص هوشمند شیت و ستون‌ها

فلو:
  1. detect_sheet()   → شناسایی شیت مناسب از workbook
  2. detect_columns() → نگاشت ستون‌ها به فیلدهای استاندارد
  3. score_mapping()  → محاسبه امتیاز کلی (0.0–1.0)

گسترش‌پذیری:
  برای اضافه کردن کسب‌وکار جدید فقط کافیست در config.py
  یک BusinessConfig با sub_categories و field aliases تعریف کنید.
  این ماژول به‌صورت خودکار تمام alias ها را از config می‌خواند.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from typing import Literal

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.business.config import BusinessConfig, SubCategory, BusinessField
from app.utils.logger import log


# ─── نوع اطمینان ─────────────────────────────────────────────────────────────
ConfidenceLevel = Literal["HIGH", "MEDIUM", "LOW", "NONE"]
DetectionMethod  = Literal["auto_exact", "auto_fuzzy", "auto_heuristic", "wizard", "none"]

# آستانه‌های امتیاز برای تصمیم‌گیری
SCORE_ACCEPT  = 0.85   # قبول کامل — بدون سوال
SCORE_CONFIRM = 0.55   # نمایش پیش‌نمایش — تأیید کاربر
# زیر 0.55 → ویزارد


# ═══════════════════════════════════════════════════════════════════════════════
# دیکشنری alias های شیت — گسترش‌پذیر
# ═══════════════════════════════════════════════════════════════════════════════

# کلید: worksheet_name تعریف‌شده در SubCategory
# مقدار: لیست alias هایی که به آن worksheet_name نگاشته می‌شوند
# برای هر SubCategory جدید، alias را اینجا یا مستقیماً در config اضافه کنید.

_SHEET_ALIAS_EXTRA: dict[str, list[str]] = {
    # پوشاک
    "clothing": [
        "clothing", "clothes", "لباس", "پوشاک", "apparel",
        "dress", "tshirt", "t-shirt", "shirt", "garment", "wear",
        "fashion", "moda", "مد", "لباس‌ها", "كلاهشان",
    ],
    # کفش
    "shoes": [
        "shoes", "footwear", "کفش", "sneakers", "boots", "sandals",
        "کفش‌ها",
    ],
    # لپتاپ
    "laptops": [
        "laptops", "laptop", "لپتاپ", "لپ‌تاپ", "notebook",
    ],
    # مانیتور
    "monitors": [
        "monitors", "monitor", "مانیتور",
    ],
    # موبایل
    "smartphones": [
        "smartphones", "smartphone", "mobile", "موبایل", "گوشی",
        "phones", "phone",
    ],
    # محصولات عمومی
    "products": [
        "products", "product", "محصولات", "محصول", "items", "item",
        "data", "sheet", "sheet1", "sheet2", "list", "catalog",
        "کالا", "اجناس",
    ],
    # قطعات
    "components": [
        "components", "component", "parts", "part", "قطعات",
    ],
    # لوازم جانبی
    "accessories": [
        "accessories", "accessory", "acc", "لوازم", "جانبی",
        "لوازم‌جانبی",
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
# خروجی‌های تشخیص
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class SheetDetectionResult:
    sheet_name: str | None          = None
    subcategory: SubCategory | None = None
    confidence: ConfidenceLevel     = "NONE"
    method: DetectionMethod         = "none"
    # تمام شیت‌های موجود در workbook (برای ویزارد)
    all_sheet_names: list[str]      = field(default_factory=list)


@dataclass
class ColumnDetectionResult:
    # فیلد_کلید → ایندکس ستون
    column_map: dict[str, int]      = field(default_factory=dict)
    # هدرهای خام شیت
    raw_headers: list[str]          = field(default_factory=list)
    # فیلدهای اجباری که پیدا نشدند
    missing_required: list[BusinessField] = field(default_factory=list)
    # امتیاز نگاشت (0.0–1.0)
    score: float                    = 0.0
    method: DetectionMethod         = "none"


@dataclass
class SmartDetectionResult:
    sheet: SheetDetectionResult     = field(default_factory=SheetDetectionResult)
    columns: ColumnDetectionResult  = field(default_factory=ColumnDetectionResult)
    # امتیاز کلی
    overall_score: float            = 0.0
    # آیا نیاز به ویزارد داریم؟
    needs_wizard: bool              = True
    # آیا باید از کاربر تأیید بگیریم؟
    needs_confirm: bool             = False


# ═══════════════════════════════════════════════════════════════════════════════
# توابع کمکی
# ═══════════════════════════════════════════════════════════════════════════════

def _normalize(text: str) -> str:
    """نرمال‌سازی: lowercase + حذف فاصله + یکنواخت‌سازی حروف فارسی/عربی"""
    if not text:
        return ""
    text = str(text).strip().lower()
    # یکنواخت‌سازی یونیکد
    text = unicodedata.normalize("NFKC", text)
    # ی/ک فارسی → استاندارد
    text = text.replace("ي", "ی").replace("ك", "ک")
    # حذف نقطه‌گذاری و خط تیره
    text = re.sub(r"[\s\-_/\\\.،,]+", "", text)
    return text


def _levenshtein(a: str, b: str) -> int:
    """فاصله لونشتاین ساده"""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(
                prev[j] + 1,
                curr[j - 1] + 1,
                prev[j - 1] + (0 if ca == cb else 1),
            ))
        prev = curr
    return prev[-1]


def _fuzzy_match(needle: str, haystack: list[str], max_dist: int = 2) -> tuple[str | None, int]:
    """پیدا کردن نزدیک‌ترین رشته در لیست با فاصله لونشتاین"""
    best, best_dist = None, max_dist + 1
    for item in haystack:
        d = _levenshtein(needle, item)
        if d < best_dist:
            best_dist = d
            best = item
    return (best, best_dist) if best_dist <= max_dist else (None, best_dist)


def _build_sheet_alias_map(business_config: BusinessConfig) -> dict[str, str]:
    """
    ساخت دیکشنری: نرمال‌سازی‌شده‌ی alias → worksheet_name
    ترکیبی از _SHEET_ALIAS_EXTRA و worksheet_name های تعریف‌شده در config.
    """
    alias_map: dict[str, str] = {}

    for sc in business_config.sub_categories:
        wn = sc.worksheet_name
        # خود نام worksheet
        alias_map[_normalize(wn)] = wn
        # alias های اضافه از دیکشنری بالا
        for alias in _SHEET_ALIAS_EXTRA.get(wn, []):
            alias_map[_normalize(alias)] = wn

    return alias_map


def _collect_field_aliases(f: BusinessField) -> list[str]:
    """تمام alias های یک فیلد را جمع می‌کند (excel_column + aliases + label)"""
    candidates = [f.excel_column]
    if hasattr(f, "aliases") and f.aliases:
        candidates.extend(f.aliases)
    if f.label_fa:
        candidates.append(f.label_fa)
    return candidates


# ═══════════════════════════════════════════════════════════════════════════════
# فاز ۱ — تشخیص شیت
# ═══════════════════════════════════════════════════════════════════════════════

def detect_sheet(
    workbook: Workbook,
    business_config: BusinessConfig,
) -> SheetDetectionResult:
    """
    شناسایی شیت مناسب از workbook.

    سه پاس به ترتیب اولویت:
      Pass 1 — تطابق دقیق alias
      Pass 2 — تطابق فازی (لونشتاین ≤ 2)
      Pass 3 — امتیاز‌دهی heuristic (شباهت ستون‌ها)
    """
    all_names = workbook.sheetnames
    result = SheetDetectionResult(all_sheet_names=all_names)

    alias_map = _build_sheet_alias_map(business_config)

    for sheet_name in all_names:
        norm = _normalize(sheet_name)

        # ─── Pass 1: exact ───
        if norm in alias_map:
            wn = alias_map[norm]
            sc = _find_subcategory_by_worksheet(business_config, wn)
            if sc:
                result.sheet_name  = sheet_name
                result.subcategory = sc
                result.confidence  = "HIGH"
                result.method      = "auto_exact"
                log.info(f"[SmartDetect] شیت '{sheet_name}' → exact → {sc.key}")
                return result

        # ─── Pass 2: fuzzy ───
        matched_alias, dist = _fuzzy_match(norm, list(alias_map.keys()), max_dist=2)
        if matched_alias:
            wn = alias_map[matched_alias]
            sc = _find_subcategory_by_worksheet(business_config, wn)
            if sc:
                result.sheet_name  = sheet_name
                result.subcategory = sc
                result.confidence  = "MEDIUM"
                result.method      = "auto_fuzzy"
                log.info(f"[SmartDetect] شیت '{sheet_name}' → fuzzy (d={dist}) → {sc.key}")
                return result

    # ─── Pass 3: heuristic ───
    best_score, best_sheet, best_sc = 0.0, None, None
    for sheet_name in all_names:
        sheet = workbook[sheet_name]
        headers = _read_headers(sheet)
        for sc in business_config.sub_categories:
            score = _score_headers_against_subcategory(headers, sc)
            if score > best_score:
                best_score, best_sheet, best_sc = score, sheet_name, sc

    if best_sc and best_score >= 0.40:
        result.sheet_name  = best_sheet
        result.subcategory = best_sc
        result.confidence  = "LOW"
        result.method      = "auto_heuristic"
        log.info(f"[SmartDetect] شیت '{best_sheet}' → heuristic (score={best_score:.2f}) → {best_sc.key}")
        return result

    log.warning("[SmartDetect] شیت مناسب پیدا نشد")
    return result  # result.sheet_name = None → ویزارد


def _find_subcategory_by_worksheet(
    business_config: BusinessConfig, worksheet_name: str
) -> SubCategory | None:
    for sc in business_config.sub_categories:
        if sc.worksheet_name.lower() == worksheet_name.lower():
            return sc
    return None


def _read_headers(sheet) -> list[str]:
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    return headers


def _score_headers_against_subcategory(
    headers: list[str], subcategory: SubCategory
) -> float:
    """امتیاز شباهت بین هدرهای یک شیت و فیلدهای یک SubCategory"""
    if not headers:
        return 0.0
    norm_headers = [_normalize(h) for h in headers]
    hits = 0
    total = len(subcategory.fields)

    for f in subcategory.fields:
        aliases = [_normalize(a) for a in _collect_field_aliases(f)]
        if any(a in norm_headers for a in aliases):
            hits += 2 if f.required else 1
        else:
            # تلاش fuzzy
            for a in aliases:
                m, _ = _fuzzy_match(a, norm_headers, max_dist=1)
                if m:
                    hits += 1
                    break

    max_score = sum(2 if f.required else 1 for f in subcategory.fields)
    return hits / max_score if max_score else 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# فاز ۲ — تشخیص ستون‌ها
# ═══════════════════════════════════════════════════════════════════════════════

def detect_columns(
    headers: list[str],
    subcategory: SubCategory,
    custom_map: dict[str, int] | None = None,
) -> ColumnDetectionResult:
    """
    نگاشت هدرهای شیت به فیلدهای استاندارد SubCategory.

    استراتژی‌ها (به ترتیب):
      1. تطابق دقیق alias
      2. تطابق فازی alias
      3. تطابق substring
    custom_map (از ویزارد) همیشه override می‌کند.
    """
    result = ColumnDetectionResult(raw_headers=headers)
    norm_headers = [_normalize(h) for h in headers]
    used_indices: set[int] = set()

    for f in subcategory.fields:
        if custom_map and f.key in custom_map:
            idx = custom_map[f.key]
            result.column_map[f.key] = idx
            used_indices.add(idx)
            continue

        aliases = [_normalize(a) for a in _collect_field_aliases(f)]

        # استراتژی ۱: exact
        found = False
        for alias in aliases:
            if alias in norm_headers:
                idx = norm_headers.index(alias)
                if idx not in used_indices:
                    result.column_map[f.key] = idx
                    used_indices.add(idx)
                    found = True
                    break

        if found:
            continue

        # استراتژی ۲: fuzzy (لونشتاین ≤ 2)
        for alias in aliases:
            matched, _ = _fuzzy_match(alias, norm_headers, max_dist=2)
            if matched:
                idx = norm_headers.index(matched)
                if idx not in used_indices:
                    result.column_map[f.key] = idx
                    used_indices.add(idx)
                    found = True
                    break

        if found:
            continue

        # استراتژی ۳: substring
        for alias in aliases:
            for idx, nh in enumerate(norm_headers):
                if (alias in nh or nh in alias) and idx not in used_indices:
                    result.column_map[f.key] = idx
                    used_indices.add(idx)
                    found = True
                    break
            if found:
                break

    # فیلدهای اجباری گمشده
    result.missing_required = [
        f for f in subcategory.fields
        if f.required and f.key not in result.column_map
    ]

    # امتیاز
    result.score = _score_column_map(result.column_map, subcategory)
    result.method = "auto_exact" if result.score >= SCORE_ACCEPT else (
        "auto_heuristic" if result.score >= SCORE_CONFIRM else "wizard"
    )

    return result


def _score_column_map(column_map: dict[str, int], subcategory: SubCategory) -> float:
    """محاسبه امتیاز نگاشت ستون‌ها"""
    score = 0
    max_score = 0
    for f in subcategory.fields:
        weight = 2 if f.required else 1
        max_score += weight
        if f.key in column_map:
            score += weight

    return score / max_score if max_score else 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# ورودی اصلی: تشخیص کامل از روی فایل
# ═══════════════════════════════════════════════════════════════════════════════

def run_smart_detection(
    file_path: str,
    business_config: BusinessConfig,
    custom_map: dict[str, int] | None = None,
) -> SmartDetectionResult:
    """
    تشخیص کامل: شیت + ستون.
    خروجی: SmartDetectionResult با needs_wizard / needs_confirm.
    """
    result = SmartDetectionResult()

    try:
        wb: Workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    except Exception as e:
        log.error(f"[SmartDetect] خطا در باز کردن فایل: {e}")
        result.needs_wizard = True
        return result

    # فاز ۱: شیت
    sheet_result = detect_sheet(wb, business_config)
    result.sheet = sheet_result

    if not sheet_result.subcategory:
        result.needs_wizard = True
        return result

    # فاز ۲: ستون‌ها
    sheet = wb[sheet_result.sheet_name]
    headers = _read_headers(sheet)
    col_result = detect_columns(headers, sheet_result.subcategory, custom_map)
    result.columns = col_result

    # امتیاز کلی
    result.overall_score = col_result.score

    if result.overall_score >= SCORE_ACCEPT and not col_result.missing_required:
        result.needs_wizard   = False
        result.needs_confirm  = False
    elif result.overall_score >= SCORE_CONFIRM:
        result.needs_wizard   = False
        result.needs_confirm  = True
    else:
        result.needs_wizard   = True
        result.needs_confirm  = False

    log.info(
        f"[SmartDetect] نتیجه: sheet={sheet_result.sheet_name}, "
        f"score={result.overall_score:.2f}, wizard={result.needs_wizard}, "
        f"confirm={result.needs_confirm}"
    )
    return result
