"""
سرویس خواندن فایل اکسل چند شیتی (با پشتیبانی از Smart Matcher و ویزارد مپینگ)
"""

from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import load_workbook
import re
import uuid
import hashlib

from app.business.config import (
    BusinessConfig,
    SubCategory,
    get_subcategory_by_worksheet,
)
from app.utils.logger import log

# 💡 فقط شیت‌های راهنما/توضیحات واقعی skip می‌شن، نه اسم‌های پیش‌فرض
# مثل "Sheet1" که خیلی از مشتری‌ها اصلاً عوضش نمی‌کنن.
IGNORED_WORKSHEET_NAMES = {"راهنما", "راهنمای استفاده", "info", "instructions", "guide", "template", "قالب"}


def generate_deterministic_sku(product_name: str, subcategory_key: str, extra_seed: str = "") -> str:
    """
    ساخت SKU پایدار (deterministic) وقتی ستون کد محصول در فایل موجود نیست.
    بر پایه‌ی نام محصول ساخته می‌شه تا در sync های بعدی همون SKU دوباره
    تولید بشه و محصول به‌جای duplicate شدن، درست آپدیت بخوره.
    ⚠️ اگه دو محصول دقیقاً اسم یکسان داشته باشن، SKU یکسان می‌گیرن (به‌عنوان یک محصول در نظر گرفته می‌شن).
    """
    normalized_name = " ".join((product_name or "").strip().lower().split())
    basis = f"{subcategory_key}|{normalized_name}|{extra_seed}"
    digest = hashlib.sha256(basis.encode("utf-8")).hexdigest()
    return f"AUTO-{digest[:10].upper()}"


# ─────────────────────────────────────────────────────────────────────────────
# فیلدهای Yes/No (ویژگی‌های حضور/عدم‌حضور یک قابلیت — Touch, Pen, LTE, ...)
# یک نقطه‌ی مرکزی برای همه‌ی فیلدهای بولی، تا هر کسب‌وکاری بخواد از این‌ها
# استفاده کنه فقط کافیه field.key رو به همین مجموعه اضافه کنه.
# ─────────────────────────────────────────────────────────────────────────────

YES_NO_FIELD_KEYS = {
    "touch_screen", "pen_support", "x360", "lte",
    "dvd_rw", "backlit_keyboard", "fingerprint", "facial_recognition",
    "hdmi", "dp", "vga_port", "lan", "thunderbolt",
}

_YES_PATTERNS = {"yes", "y", "بله", "دارد", "true", "1", "✓", "✔"}
_NO_PATTERNS = {"no", "n", "خیر", "ندارد", "false", "0", "✗", "✘", "-"}


def _parse_yes_no(value) -> str:
    """
    نرمال‌سازی مقدار به 'دارد' / 'ندارد' تا مستقیم به‌عنوان {placeholder}
    توی قالب پست قابل استفاده باشه، بدون نیاز به تبدیل جداگانه در جای دیگه.
    مقدار ناشناخته (نه yes نه no) عیناً پاس داده می‌شه — فیلد اختیاریه،
    نباید کل محصول رو به‌خاطر یه مقدار غیرمنتظره رد کنیم.
    """
    normalized = str(value).strip().lower()
    if normalized in _YES_PATTERNS:
        return "دارد"
    if normalized in _NO_PATTERNS:
        return "ندارد"
    return str(value).strip()


def _parse_optional_int(value) -> int | None:
    """
    پارس یک عدد صحیح اختیاری (مثل تعداد پورت USB یا ساعت باتری).
    اعشار به پایین رند می‌شه. مقدار نامعتبر → None (فیلد اختیاریه، خطا نمی‌دیم).
    """
    try:
        if isinstance(value, str):
            value = value.strip().replace(",", "").replace("،", "")
        return int(float(value))
    except (ValueError, TypeError):
        return None


@dataclass
class RowError:
    row_number: int
    field: str
    message: str
    worksheet: str = ""
    error_type: str = "validation"  # missing_column یا validation
    field_object: any = None       # شیء فیلد گمشده جهت استفاده در ویزارد


@dataclass
class WorksheetReadResult:
    """نتیجه خواندن یک sheet"""
    worksheet_name: str
    subcategory_key: str = ""
    headers: list[str] = field(default_factory=list)
    products: list[dict] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    total_rows: int = 0
    valid_rows: int = 0


@dataclass
class ExcelReadResult:
    """نتیجه کلی خواندن فایل"""
    worksheets: list[WorksheetReadResult] = field(default_factory=list)

    @property
    def total_products(self) -> int:
        return sum(len(ws.products) for ws in self.worksheets)

    @property
    def total_rows(self) -> int:
        return sum(ws.total_rows for ws in self.worksheets)

    @property
    def valid_rows(self) -> int:
        return sum(ws.valid_rows for ws in self.worksheets)

    @property
    def all_errors(self) -> list[RowError]:
        errors = []
        for ws in self.worksheets:
            errors.extend(ws.errors)
        return errors

    @property
    def has_errors(self) -> bool:
        return len(self.all_errors) > 0

    @property
    def is_empty(self) -> bool:
        return self.valid_rows == 0

    @property
    def all_products(self) -> list[dict]:
        products = []
        for ws in self.worksheets:
            for p in ws.products:
                p["sub_category_key"] = ws.subcategory_key
                products.append(p)
        return products

    @property
    def missing_mapping_fields(self) -> list:
        """استخراج فیلدهای گمشده برای ویزارد"""
        fields = []
        seen = set()
        for err in self.all_errors:
            if getattr(err, 'error_type', '') == "missing_column" and getattr(err, 'field_object', None):
                if err.field_object.key not in seen:
                    seen.add(err.field_object.key)
                    fields.append(err.field_object)
        return fields

    @property
    def headers(self) -> list[str]:
        """گرفتن هدرهای اولین شیت"""
        if self.worksheets and hasattr(self.worksheets[0], 'headers'):
            return self.worksheets[0].headers
        return []

def read_excel_file(
    file_path: str | Path,
    business_config: BusinessConfig,
    custom_map: dict[str, int] = None,
    ignored_fields: list[str] = None,
) -> ExcelReadResult:
    """
    خواندن فایل اکسل و تبدیل به لیست محصولات
    """
    result = ExcelReadResult()
    ignored = ignored_fields or []

    try:
        workbook = load_workbook(filename=file_path, data_only=True)
    except Exception as e:
        log.error(f"خطا در باز کردن فایل اکسل: {e}")
        ws_result = WorksheetReadResult(worksheet_name="")
        ws_result.errors.append(RowError(
            row_number=0,
            field="file",
            message=f"فایل قابل خواندن نیست: {str(e)}",
        ))
        result.worksheets.append(ws_result)
        return result

    # پردازش هر sheet
    for sheet_name in workbook.sheetnames:
        # sheet راهنما یا خالی رو نادیده بگیر
        if sheet_name.strip().lower() in IGNORED_WORKSHEET_NAMES:
            continue

        # پیدا کردن زیردسته متناظر
        subcategory = get_subcategory_by_worksheet(business_config.key, sheet_name)
        
        # انعطاف‌پذیری برای کسب‌وکار "سایر" (other)
        if not subcategory and business_config.key == "other":
            if business_config.sub_categories:
                subcategory = business_config.sub_categories[0]
                log.info(f"💡 Sheet '{sheet_name}' برای کسب‌وکار سایر به زیردسته پیش‌فرض متصل شد.")

        if not subcategory:
            log.warning(f"Sheet '{sheet_name}' متعلق به هیچ زیردسته نیست، نادیده گرفته می‌شود")
            continue

        sheet = workbook[sheet_name]
        ws_result = _read_worksheet(sheet, subcategory, custom_map, ignored)
        result.worksheets.append(ws_result)

    log.info(
        f"فایل اکسل خونده شد: {len(result.worksheets)} sheet, "
        f"{result.total_rows} سطر، {result.valid_rows} معتبر"
    )

    return result


def _read_worksheet(
    sheet,
    subcategory: SubCategory,
    custom_map: dict = None,
    ignored_fields: list = None,
) -> WorksheetReadResult:
    """خواندن یک sheet خاص با استفاده از تعریف subcategory"""
    result = WorksheetReadResult(
        worksheet_name=sheet.title,
        subcategory_key=subcategory.key,
    )

    # خواندن هدر
    headers = []
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append("")

    if not any(headers):
        return result

    ignored = ignored_fields or []

    # 💡 همیشه اول نقشه‌ی خودکار (Smart Match) ساخته می‌شود،
    # سپس در صورت وجود custom_map (پاسخ‌های ویزارد)، فقط همون فیلدها override می‌شن.
    # این‌طوری فیلدهایی که خودکار درست تشخیص داده شده بودن گم نمی‌شن.
    field_map = _build_field_map(subcategory, headers)
    if custom_map:
        field_map.update(custom_map)

    # چک فیلدهای اجباری
    missing_required = _check_missing_required(subcategory, field_map, ignored)
    if missing_required:
        for field_obj in missing_required:
            result.errors.append(RowError(
                row_number=1,
                field=field_obj.excel_column,
                message=f"ستون '{field_obj.excel_column}' در sheet '{sheet.title}' پیدا نشد (اجباری)",
                worksheet=sheet.title,
                error_type="missing_column",
                field_object=field_obj
            ))
        return result

    # خواندن ردیف‌های داده
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        result.total_rows += 1

        product_data, row_errors = _parse_row(
            row=row,
            field_map=field_map,
            subcategory=subcategory,
            row_number=row_index,
            worksheet_name=sheet.title,
            ignored_fields=ignored,
        )

        if row_errors:
            result.errors.extend(row_errors)
        else:
            result.products.append(product_data)
            result.valid_rows += 1

    return result


def _normalize_header(h) -> str:
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _header_words(h: str) -> set:
    return set(re.findall(r"[a-z0-9\u0600-\u06FF]+", h))


def _build_field_map(subcategory: SubCategory, headers: list[str]) -> dict[str, int]:
    """
    نگاشت هوشمند فیلدها با سیستم امتیازدهی سطح‌بندی‌شده:
    3 = تطبیق دقیق (کل متن ستون برابر نام فیلد/alias)
    2 = تطبیق کلمه‌به‌کلمه (کلمات alias زیرمجموعه‌ی کلمات هدر هستن)
    1 = تطبیق substring ضعیف (فقط آخرین راه‌حل)
    هیچ دو فیلدی به یک ستون مپ نمی‌شن؛ در تعارض، امتیاز بالاتر برنده و
    فیلد بازنده نامپ می‌مونه (می‌ره توی ویزارد) به‌جای حدس اشتباه.
    """
    normalized_headers = [_normalize_header(h) for h in headers]
    header_word_sets = [_header_words(h) for h in normalized_headers]

    candidates: dict[str, tuple[int, int]] = {}

    for f in subcategory.fields:
        names = [f.excel_column.strip().lower()] + [a.strip().lower() for a in (f.aliases or [])]
        best_idx, best_score = None, 0

        for name in names:
            name_words = _header_words(name)
            for idx, header in enumerate(normalized_headers):
                if header == name:
                    score = 3
                elif name_words and name_words.issubset(header_word_sets[idx]):
                    score = 2
                elif name in header or header in name:
                    score = 1
                else:
                    continue
                if score > best_score:
                    best_score, best_idx = score, idx

        if best_idx is not None:
            candidates[f.key] = (best_idx, best_score)

    field_map: dict[str, int] = {}
    col_owner: dict[int, tuple[str, int]] = {}

    for field_key, (col_idx, score) in candidates.items():
        owner = col_owner.get(col_idx)
        if owner is None or score > owner[1]:
            if owner is not None:
                field_map.pop(owner[0], None)
            col_owner[col_idx] = (field_key, score)
            field_map[field_key] = col_idx

    return field_map

def _check_missing_required(subcategory: SubCategory, field_map: dict, ignored_fields: list) -> list:
    """چک کن فیلدهای اجباری جا نمونده باشن (خروجی: لیست شیء فیلدها)"""
    missing = []
    ignored = ignored_fields or []
    for field in subcategory.fields:
        if field.required and field.key not in field_map and field.key not in ignored:
            missing.append(field)
    return missing

def _parse_row(
    row: tuple,
    field_map: dict[str, int],
    subcategory: SubCategory,
    row_number: int,
    worksheet_name: str,
    ignored_fields: list = None,
) -> tuple[dict, list[RowError]]:
    """پارس یک ردیف"""
    product_data = {"row_number": row_number, "specs": {}}
    errors = []
    ignored = ignored_fields or []

    # مرحله ۱: پردازش فیلدهای معمولی (غیر ignored) — از جمله product_name
    for field in subcategory.fields:
        if field.key in ignored:
            continue

        if field.key not in field_map:
            continue

        col_index = field_map[field.key]
        raw_value = row[col_index] if col_index < len(row) else None
        value = _clean_value(raw_value)

        if field.required and (value is None or value == ""):
            errors.append(RowError(
                row_number=row_number,
                field=field.excel_column,
                message=f"مقدار '{field.excel_column}' خالی است",
                worksheet=worksheet_name,
            ))
            continue

        parsed_value, parse_error = _parse_field_value(field.key, value, field.excel_column)

        if parse_error:
            errors.append(RowError(
                row_number=row_number,
                field=field.excel_column,
                message=parse_error,
                worksheet=worksheet_name,
            ))
            continue

        if field.key in ("sku", "product_name", "price", "stock", "description", "image_url"):
            product_data[field.key] = parsed_value
        else:
            # 🆕 مقدار خالی/None اصلاً در specs ذخیره نمی‌شه — یعنی این ویژگی
            # برای این محصول اصلاً وجود نداره و بعداً در قالب پست ذکر نمی‌شه
            if parsed_value not in (None, ""):
                product_data["specs"][field.key] = parsed_value

    # مرحله ۲: تولید مقدار خودکار برای فیلدهای ignored
    # (product_name اول ساخته می‌شه چون sku خودکار به اون وابسته‌ست)
    if "product_name" in ignored:
        product_data["product_name"] = f"محصول بدون نام - ردیف {row_number}"

    for field in subcategory.fields:
        if field.key not in ignored or field.key == "product_name":
            continue

        if field.key == "sku":
            parsed_value = generate_deterministic_sku(
                product_name=product_data.get("product_name", ""),
                subcategory_key=subcategory.key,
                extra_seed=worksheet_name,
            )
        elif field.key == "price":
            parsed_value = 0
        elif field.key == "stock":
            parsed_value = 0
        else:
            parsed_value = ""

        if field.key in ("sku", "product_name", "price", "stock", "description", "image_url"):
            product_data[field.key] = parsed_value
        else:
            if parsed_value not in (None, ""):
                product_data["specs"][field.key] = parsed_value

    return product_data, errors


def _clean_value(value):
    """پاکسازی مقدار سلول"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def _parse_field_value(field_key: str, value, column_name: str) -> tuple:
    """اعتبارسنجی و تبدیل نوع"""
    if value is None or value == "":
        return None, None

    if field_key == "price":
        try:
            if isinstance(value, str):
                value = value.replace(",", "").replace("،", "").strip()
            price = int(float(value))
            if price < 0:
                return None, "قیمت نمی‌تواند منفی باشد"
            return price, None
        except (ValueError, TypeError):
            return None, f"قیمت باید عدد باشد (مقدار: {value})"

    if field_key == "stock":
        try:
            stock = int(float(value))
            if stock < 0:
                return None, "موجودی نمی‌تواند منفی باشد"
            return stock, None
        except (ValueError, TypeError):
            return None, f"موجودی باید عدد باشد (مقدار: {value})"

    if field_key == "sku":
        sku = str(value).strip()
        if not sku:
            return None, "کد محصول نمی‌تواند خالی باشد"
        if len(sku) > 80:
            return None, "کد محصول نباید بیش از ۸۰ کاراکتر باشد"
        return sku, None

    if field_key == "product_name":
        name = str(value).strip()
        if not name:
            return None, "نام محصول نمی‌تواند خالی باشد"
        if len(name) > 250:
            return None, "نام محصول نباید بیش از ۲۵۰ کاراکتر باشد"
        return name, None

    # 🆕 فیلدهای Yes/No (Touch, Pen, LTE, HDMI, DP, VP, LAN, ...)
    if field_key in YES_NO_FIELD_KEYS:
        return _parse_yes_no(value), None

    # 🆕 تعداد پورت USB
    if field_key == "usb_ports":
        return _parse_optional_int(value), None

    # 🆕 عمر باتری (ساعت) — اعشار به پایین رند می‌شه
    if field_key == "battery_life":
        return _parse_optional_int(value), None

    # Grade و Weight و بقیه‌ی فیلدهای متنی: عیناً پاس داده می‌شن (بدون تغییر)
    return str(value).strip(), None