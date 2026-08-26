"""
ویزارد هوشمند مپینگ ۵ مرحله‌ای

مرحله ۱: انتخاب شیت (اگر auto-detect ناموفق بود)
مرحله ۲: انتخاب زیردسته / نوع محصول
مرحله ۳: نگاشت ستون‌های اجباری (یکی یکی)
مرحله ۴: نگاشت ستون‌های اختیاری
مرحله ۵: بررسی و تأیید نهایی → ذخیره

این ویزارد جایگزین mapping_wizard.py موجود نمی‌شود؛
آن برای شیت‌های Google Sheet بدون شناسایی اجرا می‌شود.
این ویزارد برای اکسل آپلودی طراحی شده است.
"""

import os
import tempfile

from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ContextTypes

from app.bot.states.user_state import (
    UserState,
    set_user_state,
    get_user_state,
    get_user_data,
    clear_user_state,
)
from app.business.config import get_business, get_all_businesses
from app.services.data_input.smart_detector import (
    detect_columns,
    run_smart_detection,
    SCORE_ACCEPT,
    SCORE_CONFIRM,
)
from app.services.mapping_service import save_mapping_from_wizard
from app.services.data_input.excel_reader import read_excel_file
from app.services.product_service import save_products_from_excel
from app.services.customer_service import get_customer_by_telegram_id
from app.services.business_service import (
    get_business_config_for_customer,
    get_business_for_customer,
)
from app.services.subscription.service import get_active_subscription
from app.services.subscription.plans import get_plan
from app.database.connection import AsyncSessionLocal
from app.utils.logger import log

# مرحله‌های ویزارد
STEP_SHEET        = "sheet"
STEP_SUBTYPE      = "subtype"
STEP_REQUIRED_COL = "req_col"
STEP_OPTIONAL_COL = "opt_col"
STEP_REVIEW       = "review"


# ─────────────────────────────────────────────────────────────────────────────
# ورودی ویزارد
# ─────────────────────────────────────────────────────────────────────────────

async def start_smart_wizard(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    user_id: int,
    file_path: str,
    business_type_key: str,
    all_sheet_names: list[str],
    detected_sheet: str | None = None,
    detected_subcategory_key: str | None = None,
):
    """
    شروع ویزارد هوشمند.
    اگر شیت قبلاً شناخته شده باشد (detected_sheet) مستقیم به مرحله ۲ می‌رود.
    """
    set_user_state(
        user_id,
        UserState.SMART_MAPPING_WIZARD,
        data={
            "file_path":              file_path,
            "business_type_key":      business_type_key,
            "all_sheet_names":        all_sheet_names,
            "selected_sheet":         detected_sheet,
            "selected_subcategory":   detected_subcategory_key,
            "column_map":             {},
            "ignored_fields":         [],
            "required_fields_queue":  [],
            "optional_fields_queue":  [],
            "step":                   STEP_SHEET if not detected_sheet else STEP_SUBTYPE,
        },
    )

    if detected_sheet and detected_subcategory_key:
        await _go_to_required_cols(update, user_id)
    elif detected_sheet:
        await _ask_subtype(update, user_id, business_type_key)
    else:
        await _ask_sheet(update, user_id)


# ─────────────────────────────────────────────────────────────────────────────
# مرحله ۱ — انتخاب شیت
# ─────────────────────────────────────────────────────────────────────────────

async def _ask_sheet(update: Update, user_id: int):
    user_data = get_user_data(user_id)
    sheet_names = user_data.get("all_sheet_names", [])

    keyboard = []
    for name in sheet_names:
        keyboard.append([InlineKeyboardButton(
            f"📄 {name}", callback_data=f"smwiz_sheet_{name[:40]}"
        )])
    keyboard.append([InlineKeyboardButton("🚫 لغو", callback_data="smwiz_cancel")])

    text = (
        "📋 <b>مرحله ۱ از ۵ — انتخاب شیت</b>\n"
        "━━━━━━━━━━━━━━━\n\n"
        "شیت مورد نظر (حاوی اطلاعات محصولات) را انتخاب کنید:"
    )

    msg = update.callback_query.message if update.callback_query else update.message
    await msg.reply_text(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(keyboard))


async def wizard_sheet_selected_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if get_user_state(user_id) != UserState.SMART_MAPPING_WIZARD:
        return

    sheet_name = query.data.replace("smwiz_sheet_", "")
    user_data  = get_user_data(user_id)
    user_data["selected_sheet"] = sheet_name
    user_data["step"] = STEP_SUBTYPE
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)

    await _ask_subtype(update, user_id, user_data.get("business_type_key", ""))


# ─────────────────────────────────────────────────────────────────────────────
# مرحله ۲ — تأیید / انتخاب زیردسته
# ─────────────────────────────────────────────────────────────────────────────

async def _ask_subtype(update: Update, user_id: int, business_type_key: str):
    business = get_business(business_type_key)
    if not business:
        return

    keyboard = []
    for sc in business.sub_categories:
        keyboard.append([InlineKeyboardButton(
            f"{sc.emoji} {sc.name_fa}",
            callback_data=f"smwiz_sub_{sc.key}"
        )])
    keyboard.append([InlineKeyboardButton("🚫 لغو", callback_data="smwiz_cancel")])

    text = (
        "🏷 <b>مرحله ۲ از ۵ — نوع محصولات</b>\n"
        "━━━━━━━━━━━━━━━\n\n"
        "محصولات این شیت در چه دسته‌ای قرار می‌گیرند؟"
    )

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await update.message.reply_text(
            text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def wizard_subtype_selected_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if get_user_state(user_id) != UserState.SMART_MAPPING_WIZARD:
        return

    subcategory_key = query.data.replace("smwiz_sub_", "")
    user_data = get_user_data(user_id)
    user_data["selected_subcategory"] = subcategory_key
    user_data["step"] = STEP_REQUIRED_COL
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)

    await _go_to_required_cols(update, user_id)


async def _go_to_required_cols(update: Update, user_id: int):
    """آماده‌سازی صف ستون‌های اجباری و اختیاری"""
    user_data = get_user_data(user_id)
    business_type_key = user_data.get("business_type_key", "")
    subcategory_key   = user_data.get("selected_subcategory", "")
    file_path         = user_data.get("file_path", "")
    sheet_name        = user_data.get("selected_sheet", "")

    business = get_business(business_type_key)
    if not business:
        return

    subcategory = next(
        (sc for sc in business.sub_categories if sc.key == subcategory_key),
        business.sub_categories[0] if business.sub_categories else None
    )
    if not subcategory:
        return

    # خواندن هدرهای شیت از فایل
    from openpyxl import load_workbook
    try:
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        raw_headers = [
            str(cell.value).strip()
            for cell in sheet[1]
            if cell.value
        ]
    except Exception as e:
        log.error(f"[SmartWizard] خطا در خواندن هدرها: {e}")
        raw_headers = []

    # تشخیص اولیه ستون‌ها
    existing_map = detect_columns(raw_headers, subcategory)

    # ستون‌های اجباری و اختیاری که هنوز نگاشته نشده‌اند
    required_queue = [
        f for f in subcategory.fields
        if f.required and f.key not in existing_map.column_map
    ]
    optional_queue = [
        f for f in subcategory.fields
        if not f.required and f.key not in existing_map.column_map
    ]

    user_data.update({
        "raw_headers":           raw_headers,
        "column_map":            existing_map.column_map,
        "required_fields_queue": required_queue,
        "optional_fields_queue": optional_queue,
        "subcategory_obj":       None,  # نمی‌توان ORM object ذخیره کرد
        "step":                  STEP_REQUIRED_COL,
    })
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)

    await _ask_next_required_col(update, user_id)


# ─────────────────────────────────────────────────────────────────────────────
# مرحله ۳ — ستون‌های اجباری
# ─────────────────────────────────────────────────────────────────────────────

async def _ask_next_required_col(update: Update, user_id: int):
    user_data = get_user_data(user_id)
    queue     = user_data.get("required_fields_queue", [])
    headers   = user_data.get("raw_headers", [])

    if not queue:
        # همه اجباری‌ها تمام شد — برو به اختیاری‌ها
        await _ask_next_optional_col(update, user_id)
        return

    field = queue[0]
    total_req = len(user_data.get("required_fields_queue", [])) + len(
        [f for f in user_data.get("column_map", {}).keys()]
    )

    text = (
        f"🔴 <b>مرحله ۳ از ۵ — ستون اجباری</b>\n"
        f"━━━━━━━━━━━━━━━\n\n"
        f"ستون مربوط به <b>«{field.emoji} {field.label_fa}»</b> کدام است؟\n\n"
        f"<i>این فیلد اجباری است و حذف آن باعث رد شدن ردیف‌ها می‌شود.</i>"
    )

    await _send_column_keyboard(update, headers, text, show_ignore=False)


async def wizard_req_col_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """کاربر یک ستون اجباری انتخاب کرد"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if get_user_state(user_id) != UserState.SMART_MAPPING_WIZARD:
        return

    data      = query.data
    user_data = get_user_data(user_id)
    queue     = user_data.get("required_fields_queue", [])

    if not queue:
        await _ask_next_optional_col(update, user_id)
        return

    current_field = queue.pop(0)

    if data.startswith("smwiz_col_"):
        col_index = int(data.replace("smwiz_col_", ""))
        user_data["column_map"][current_field.key] = col_index

    user_data["required_fields_queue"] = queue
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)

    await _ask_next_required_col(update, user_id)


# ─────────────────────────────────────────────────────────────────────────────
# مرحله ۴ — ستون‌های اختیاری
# ─────────────────────────────────────────────────────────────────────────────

async def _ask_next_optional_col(update: Update, user_id: int):
    user_data = get_user_data(user_id)
    queue     = user_data.get("optional_fields_queue", [])
    headers   = user_data.get("raw_headers", [])

    if not queue:
        # همه تمام شد — برو به مرور
        await _show_review(update, user_id)
        return

    field = queue[0]
    text = (
        f"⬜ <b>مرحله ۴ از ۵ — ستون اختیاری</b>\n"
        f"━━━━━━━━━━━━━━━\n\n"
        f"ستون مربوط به <b>«{field.emoji} {field.label_fa}»</b> کدام است؟\n\n"
        f"<i>این فیلد اختیاری است. می‌توانید از آن صرف‌نظر کنید.</i>"
    )

    await _send_column_keyboard(update, headers, text, show_ignore=True)


async def wizard_opt_col_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """کاربر یک ستون اختیاری انتخاب کرد یا رد کرد"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if get_user_state(user_id) != UserState.SMART_MAPPING_WIZARD:
        return

    data      = query.data
    user_data = get_user_data(user_id)
    queue     = user_data.get("optional_fields_queue", [])

    if not queue:
        await _show_review(update, user_id)
        return

    current_field = queue.pop(0)

    if data.startswith("smwiz_col_"):
        col_index = int(data.replace("smwiz_col_", ""))
        user_data["column_map"][current_field.key] = col_index
    elif data == "smwiz_col_ignore":
        if "ignored_fields" not in user_data:
            user_data["ignored_fields"] = []
        user_data["ignored_fields"].append(current_field.key)

    user_data["optional_fields_queue"] = queue
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)

    await _ask_next_optional_col(update, user_id)


# ─────────────────────────────────────────────────────────────────────────────
# مرحله ۵ — بررسی و تأیید
# ─────────────────────────────────────────────────────────────────────────────

async def _show_review(update: Update, user_id: int):
    user_data     = get_user_data(user_id)
    column_map    = user_data.get("column_map", {})
    raw_headers   = user_data.get("raw_headers", [])
    ignored       = user_data.get("ignored_fields", [])
    biz_key       = user_data.get("business_type_key", "")
    sub_key       = user_data.get("selected_subcategory", "")

    business = get_business(biz_key)
    subcat   = next(
        (sc for sc in (business.sub_categories if business else []) if sc.key == sub_key),
        None
    )

    lines = [
        "✅ <b>مرحله ۵ از ۵ — بررسی مپینگ</b>",
        "━━━━━━━━━━━━━━━",
        "",
    ]

    if subcat:
        for f in subcat.fields:
            if f.key in column_map:
                col_idx  = column_map[f.key]
                col_name = raw_headers[col_idx] if col_idx < len(raw_headers) else f"ستون {col_idx}"
                status   = "✅"
            elif f.key in ignored:
                col_name = "رد شد"
                status   = "⬜"
            else:
                col_name = "یافت نشد"
                status   = "❓"
            req_mark = " 🔴" if f.required else ""
            lines.append(f"{status} <b>{f.label_fa}</b>{req_mark}: {col_name}")

    lines += [
        "",
        "━━━━━━━━━━━━━━━",
        "آیا این نگاشت درست است؟",
    ]

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ ذخیره و ادامه", callback_data="smwiz_confirm"),
            InlineKeyboardButton("🔄 شروع مجدد",    callback_data="smwiz_restart"),
        ],
        [InlineKeyboardButton("🚫 لغو", callback_data="smwiz_cancel")],
    ])

    if update.callback_query:
        await update.callback_query.edit_message_text(
            "\n".join(lines), parse_mode="HTML", reply_markup=keyboard
        )
    else:
        await update.message.reply_text(
            "\n".join(lines), parse_mode="HTML", reply_markup=keyboard
        )


async def wizard_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """کاربر مپینگ را تأیید کرد — ذخیره و پردازش فایل"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    if get_user_state(user_id) != UserState.SMART_MAPPING_WIZARD:
        return

    await query.edit_message_text("💾 در حال ذخیره مپینگ و پردازش فایل...")

    user_data      = get_user_data(user_id)
    file_path      = user_data.get("file_path", "")
    biz_key        = user_data.get("business_type_key", "")
    sub_key        = user_data.get("selected_subcategory", "")
    sheet_name     = user_data.get("selected_sheet", "")
    column_map     = user_data.get("column_map", {})
    ignored_fields = user_data.get("ignored_fields", [])
    raw_headers    = user_data.get("raw_headers", [])

    try:
        async with AsyncSessionLocal() as session:
            customer = await get_customer_by_telegram_id(session, user_id)
            if not customer:
                await query.edit_message_text("❌ مشتری پیدا نشد!")
                clear_user_state(user_id)
                return

            # ذخیره مپینگ در دیتابیس
            await save_mapping_from_wizard(
                session=session,
                customer_id=customer.id,
                sheet_name=sheet_name,
                subcategory_key=sub_key,
                column_map=column_map,
                ignored_fields=ignored_fields,
                raw_headers=raw_headers,
            )

            business_config = get_business_config_for_customer(customer)
            business        = await get_business_for_customer(session, customer.id)
            subscription    = await get_active_subscription(session, customer.id)

            if not business_config or not subscription:
                await query.edit_message_text("❌ اطلاعات کسب‌وکار یا اشتراک یافت نشد.")
                clear_user_state(user_id)
                return

            plan        = get_plan(subscription.plan_key)
            business_id = business.id if business else None

            # خواندن فایل با مپینگ سفارشی
            read_result = read_excel_file(
                file_path=file_path,
                business_config=business_config,
                custom_map=column_map,
                ignored_fields=ignored_fields,
            )

            save_result = await save_products_from_excel(
                session=session,
                customer_id=customer.id,
                business_id=business_id,
                products_data=read_result.all_products,
                max_products_limit=plan.max_products,
            )

        await query.edit_message_text(
            f"🎉 <b>پردازش با موفقیت انجام شد!</b>\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💾 مپینگ ذخیره شد (دفعه بعد نیازی به تنظیم مجدد ندارید)\n\n"
            f"🆕 محصولات جدید: {save_result.new_count}\n"
            f"🔄 آپدیت شده: {save_result.updated_count}\n"
            f"✅ بدون تغییر: {save_result.unchanged_count}\n"
            f"❌ خطا: {save_result.error_count}\n"
            f"━━━━━━━━━━━━━━━",
            parse_mode="HTML",
        )

    except Exception as e:
        log.error(f"[SmartWizard] خطا در پردازش نهایی: {e}", exc_info=True)
        await query.edit_message_text(f"❌ خطا: {str(e)[:200]}")

    finally:
        clear_user_state(user_id)
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except Exception:
                pass


async def wizard_restart_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """شروع مجدد ویزارد"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    user_data = get_user_data(user_id)
    # ریست صف‌ها، نگه‌داشتن فایل و biz_key
    user_data.update({
        "selected_sheet":        None,
        "selected_subcategory":  None,
        "column_map":            {},
        "ignored_fields":        [],
        "required_fields_queue": [],
        "optional_fields_queue": [],
        "step":                  STEP_SHEET,
    })
    set_user_state(user_id, UserState.SMART_MAPPING_WIZARD, data=user_data)
    await _ask_sheet(update, user_id)


async def wizard_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """لغو ویزارد"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id

    user_data = get_user_data(user_id)
    file_path = user_data.get("file_path")
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass

    clear_user_state(user_id)
    await query.edit_message_text(
        "❌ عملیات لغو شد.\n\n"
        "می‌توانید فایل جدیدی آپلود کنید."
    )


# ─────────────────────────────────────────────────────────────────────────────
# کمکی: کیبورد ستون‌ها
# ─────────────────────────────────────────────────────────────────────────────

async def _send_column_keyboard(
    update: Update,
    headers: list[str],
    text: str,
    show_ignore: bool = True,
):
    keyboard = []
    row = []
    for idx, header in enumerate(headers):
        btn_text = str(header)[:22] if header else f"ستون {idx + 1}"
        row.append(InlineKeyboardButton(btn_text, callback_data=f"smwiz_col_{idx}"))
        if len(row) == 2:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)

    if show_ignore:
        keyboard.append([InlineKeyboardButton(
            "⬜ ندارم / رد کن", callback_data="smwiz_col_ignore"
        )])
    keyboard.append([InlineKeyboardButton("🚫 لغو", callback_data="smwiz_cancel")])

    markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(
            text, parse_mode="HTML", reply_markup=markup
        )
    else:
        await update.message.reply_text(text, parse_mode="HTML", reply_markup=markup)
