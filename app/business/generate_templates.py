"""
تولید فایل‌های نمونه اکسل برای هر کسب‌وکار
هر فایل شامل چندین Sheet هست (یکی برای هر زیردسته)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.business.config import (
    get_all_businesses,
    BUSINESS_DIR,
    BusinessConfig,
    SubCategory,
)


# ─── استایل‌ها ───
HEADER_FONT = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E86DE", end_color="2E86DE", fill_type="solid")
DATA_FONT = Font(name="Tahoma", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ─── داده‌های نمونه برای هر زیردسته ───
SAMPLE_DATA = {
    "laptop": [
        ["LP001", "IdeaPad 5 Pro", "Lenovo", "Core i7-12700H", "16GB", "512GB SSD", "RTX 3050 4GB", "16 اینچ 2.5K", 42500000, 5, "لپتاپی قدرتمند برای کار و بازی", ""],
        ["LP002", "ROG Strix G16", "ASUS", "Core i9-13900H", "32GB", "1TB SSD", "RTX 4060 8GB", "16 اینچ FHD 165Hz", 65000000, 3, "لپتاپ گیمینگ حرفه‌ای", ""],
    ],
    "prebuilt_pc": [
        ["PC001", "Gaming PC Pro", "Custom", "Core i7-13700K", "32GB DDR5", "1TB NVMe", "RTX 4070", "850W Gold", "NZXT H510", 55000000, 2, "کیس گیمینگ آماده رده بالا", ""],
        ["PC002", "Office PC", "Custom", "Core i5-12400", "16GB DDR4", "512GB SSD", "UHD 730", "600W Bronze", "Cooler Master", 22000000, 5, "کیس مناسب کار اداری", ""],
    ],
    "monitor": [
        ["MN001", "Odyssey G7", "Samsung", "32 اینچ", "2K QHD", "240Hz", "VA", 28000000, 4, "مانیتور گیمینگ خمیده", ""],
        ["MN002", "ProArt Display", "ASUS", "27 اینچ", "4K UHD", "60Hz", "IPS", 35000000, 2, "مانیتور حرفه‌ای طراحی", ""],
    ],
    "component": [
        ["CM001", "Fury Beast", "Kingston", "رم", "16GB DDR4 3200MHz", 2800000, 20, "رم گیمینگ با کیفیت", ""],
        ["CM002", "980 PRO", "Samsung", "SSD", "1TB NVMe PCIe 4.0", 4500000, 10, "SSD فوق سریع", ""],
        ["CM003", "Ryzen 7 7800X3D", "AMD", "پردازنده", "8 هسته 16 نخ 5GHz", 18500000, 3, "پردازنده گیمینگ رده بالا", ""],
    ],
    "accessory": [
        ["AC001", "MX Master 3S", "Logitech", "ماوس", "بی‌سیم Bluetooth", "دقت 8000 DPI، ۷ دکمه", 4500000, 15, "ماوس حرفه‌ای", ""],
        ["AC002", "K95 RGB", "Corsair", "کیبورد", "USB", "مکانیکی، RGB، Cherry MX", 6800000, 8, "کیبورد گیمینگ مکانیکی", ""],
        ["AC003", "HD 660S", "Sennheiser", "هدفون", "3.5mm", "استودیویی، بی‌سیم", 12000000, 5, "هدفون هایفای", ""],
    ],
    "general_item": [
        ["OT001", "محصول تستی شماره یک", "برند الف", 1500000, 10, "توضیحات کوتاه درباره محصول", ""],
        ["OT002", "کالای آزمایشی ویژه", "بدون برند", 250000, 50, "کالای عالی برای تست", ""],
    ]
}


def create_worksheet_from_subcategory(wb: Workbook, subcategory: SubCategory):
    """ساخت یک sheet از یک SubCategory"""

    # ساخت sheet جدید
    ws = wb.create_sheet(title=subcategory.worksheet_name)
    ws.sheet_view.rightToLeft = True

    # هدرها از روی فیلدها
    headers = [field.excel_column for field in subcategory.fields]

    # نوشتن هدر
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER

    # داده‌های نمونه
    sample_rows = SAMPLE_DATA.get(subcategory.key, [])
    for row_num, row_data in enumerate(sample_rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = DATA_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

    # تنظیم عرض ستون‌ها
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    # ارتفاع هدر
    ws.row_dimensions[1].height = 35


def create_business_template(business: BusinessConfig):
    """ساخت فایل نمونه برای یک کسب‌وکار"""

    wb = Workbook()

    # حذف sheet پیش‌فرض
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ساخت راهنما در Sheet اول
    _create_info_sheet(wb, business)

    # ساخت یک sheet برای هر زیردسته
    for subcategory in business.sub_categories:
        create_worksheet_from_subcategory(wb, subcategory)

    # ذخیره فایل
    template_dir = BUSINESS_DIR / "templates"
    template_dir.mkdir(exist_ok=True)

    file_path = template_dir / f"{business.key}.xlsx"
    wb.save(file_path)
    print(f"✅ فایل ساخته شد: {file_path}")


def _create_info_sheet(wb: Workbook, business: BusinessConfig):
    """ساخت sheet راهنما در ابتدای فایل"""
    ws = wb.create_sheet(title="راهنما", index=0)
    ws.sheet_view.rightToLeft = True

    info_lines = [
        f"📘 راهنمای فایل نمونه: {business.name_fa}",
        "",
        "این فایل شامل چندین صفحه است:",
        "هر صفحه مخصوص یک دسته از محصولات شماست.",
        "",
        "📋 دسته‌های موجود:",
    ]

    for sc in business.sub_categories:
        info_lines.append(f"   • {sc.emoji} {sc.name_fa}  (صفحه: {sc.worksheet_name})")

    info_lines.extend([
        "",
        "⚠️ نکات مهم:",
        "1. نام صفحه‌ها را تغییر ندهید",
        "2. ردیف اول (نام ستون‌ها) را تغییر ندهید",
        "3. کد محصول باید یکتا باشد",
        "4. فقط محصولاتی که دارید را وارد کنید",
        "5. صفحه‌های خالی مشکلی ندارد (نادیده گرفته می‌شوند)",
        "",
        "💡 اگر یک دسته را ندارید، آن صفحه را خالی بگذارید",
    ])

    for row_num, line in enumerate(info_lines, 1):
        cell = ws.cell(row=row_num, column=1, value=line)
        if row_num == 1:
            cell.font = Font(name="Tahoma", size=14, bold=True, color="2E86DE")
        else:
            cell.font = Font(name="Tahoma", size=11)

    ws.column_dimensions["A"].width = 80


def create_post_templates(business: BusinessConfig):
    """ساخت قالب‌های پست برای هر زیردسته"""

    # قالب لپتاپ
    laptop_template = """💻 {product_name}

🏭 برند: {brand}
⚡ پردازنده: {cpu}
🧠 رم: {ram}
💾 حافظه: {storage}
🎮 گرافیک: {gpu}
📐 صفحه: {screen}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب کیس آماده
    prebuilt_template = """🖥 {product_name}

🏭 برند: {brand}
⚡ پردازنده: {cpu}
🧠 رم: {ram}
💾 حافظه: {storage}
🎮 گرافیک: {gpu}
🔌 منبع تغذیه: {psu}
📦 کیس: {case_model}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب مانیتور
    monitor_template = """🖥 {product_name}

🏭 برند: {brand}
📐 سایز: {screen_size}
🔍 رزولوشن: {resolution}
⚡ نرخ رفرش: {refresh_rate}
🎨 نوع پنل: {panel_type}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب قطعات
    component_template = """⚙️ {product_name}

🔧 نوع قطعه: {component_type}
🏭 برند: {brand}
📋 مشخصات: {specs}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب لوازم جانبی
    accessory_template = """🎧 {product_name}

🏭 برند: {brand}
🔧 نوع: {accessory_type}
🔌 اتصال: {connection}
✨ ویژگی‌ها: {features}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب گوشی موبایل
    smartphone_template = """📱 {product_name}

🏭 برند: {brand}
🧠 حافظه رم: {ram}
💾 حافظه داخلی: {storage}
📸 دوربین: {camera}
🔋 باتری: {battery}
🎨 رنگ: {color}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    # قالب پوشاک
    clothing_template = """👗 {product_name}

🏭 برند: {brand}
📏 سایز: {size}
🎨 رنگ‌بندی: {color}
🧶 جنس: {material}
─────────────────
💰 قیمت: {price} تومان
📦 {stock_status}
─────────────────
{description_block}

{hashtags}

📞 {contact}
🔄 {update_date}"""

    templates_map = {
        "laptop": laptop_template,
        "prebuilt_pc": prebuilt_template,
        "monitor": monitor_template,
        "component": component_template,
        "accessory": accessory_template,
        "smartphone": smartphone_template, 
        "clothing": clothing_template,
    }

    for subcategory in business.sub_categories:
        template_content = templates_map.get(subcategory.key)
        if not template_content:
            continue

        template_path = BUSINESS_DIR / subcategory.post_template_file
        template_path.parent.mkdir(parents=True, exist_ok=True)
        template_path.write_text(template_content, encoding="utf-8")
        print(f"✅ قالب ساخته شد: {template_path}")


def main():
    """تولید همه فایل‌ها"""
    print("🔨 در حال تولید فایل‌های نمونه...\n")

    for business in get_all_businesses():
        print(f"📊 کسب‌وکار: {business.name_fa}")
        create_business_template(business)
        create_post_templates(business)
        print()

    print("✅ همه فایل‌ها ساخته شدند!")


if __name__ == "__main__":
    main()