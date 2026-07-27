"""
تست خواندن فایل اکسل
"""

import pytest
import tempfile
from pathlib import Path
from openpyxl import Workbook

from app.services.data_input.excel_reader import (
    read_excel_file,
    ExcelReadResult,
    RowError,
)
from app.business.config import get_business


def create_test_excel(rows: list[list], headers: list[str]) -> str:
    """ساخت فایل اکسل تست"""
    wb = Workbook()
    ws = wb.active

    # هدر
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # داده‌ها
    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # ذخیره در فایل موقت
    temp_file = tempfile.NamedTemporaryFile(
        suffix='.xlsx', delete=False
    )
    wb.save(temp_file.name)
    temp_file.close()
    return temp_file.name


def test_read_empty_file():
    """فایل بدون داده"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    file_path = create_test_excel(rows=[], headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)
        assert result.is_empty
        assert result.valid_rows == 0
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_read_valid_file():
    """فایل با ۲ محصول معتبر"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["LP001", "IdeaPad 5", "Lenovo", "i7-12700H", "16GB", "512GB SSD", 42500000, 5],
        ["LP002", "ROG Strix", "ASUS", "i9-13900H", "32GB", "1TB SSD", 65000000, 3],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)

        assert result.valid_rows == 2
        assert result.total_rows == 2
        assert len(result.products) == 2
        assert not result.has_errors

        # چک محصول اول
        product1 = result.products[0]
        assert product1["sku"] == "LP001"
        assert product1["product_name"] == "IdeaPad 5"
        assert product1["price"] == 42500000
        assert product1["stock"] == 5
        assert product1["specs"]["brand"] == "Lenovo"
        assert product1["specs"]["cpu"] == "i7-12700H"
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_read_file_with_invalid_price():
    """فایل با قیمت غیرعددی"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["LP001", "IdeaPad", "Lenovo", "i7", "16GB", "512GB", "قیمت اشتباه", 5],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)

        assert result.has_errors
        assert result.valid_rows == 0
        # چک کن خطای قیمت هست
        assert any("قیمت" in err.field for err in result.errors)
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_read_file_missing_required_column():
    """فایل بدون ستون اجباری (SKU)"""
    laptop_business = get_business("laptop_store")
    # بدون "کد محصول"
    headers = ["نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["IdeaPad", "Lenovo", "i7", "16GB", "512GB", 42500000, 5],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)

        assert result.has_errors
        # چک کن خطای ستون اجباری هست
        assert any("کد محصول" in err.message for err in result.errors)
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_price_with_comma():
    """قیمت با کاما (42,500,000)"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["LP001", "IdeaPad", "Lenovo", "i7", "16GB", "512GB", "42,500,000", 5],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)
        assert result.valid_rows == 1
        assert result.products[0]["price"] == 42500000
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_zero_stock_valid():
    """موجودی صفر باید معتبر باشه (ناموجود)"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["LP001", "IdeaPad", "Lenovo", "i7", "16GB", "512GB", 42500000, 0],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)
        assert result.valid_rows == 1
        assert result.products[0]["stock"] == 0
    finally:
        Path(file_path).unlink(missing_ok=True)


def test_empty_rows_ignored():
    """ردیف‌های خالی نادیده گرفته بشن"""
    laptop_business = get_business("laptop_store")
    headers = ["کد محصول", "نام محصول", "برند", "پردازنده", "رم", "حافظه", "قیمت", "موجودی"]

    rows = [
        ["LP001", "IdeaPad", "Lenovo", "i7", "16GB", "512GB", 42500000, 5],
        [None, None, None, None, None, None, None, None],  # خالی
        ["LP002", "ROG", "ASUS", "i9", "32GB", "1TB", 65000000, 3],
    ]

    file_path = create_test_excel(rows=rows, headers=headers)

    try:
        result = read_excel_file(file_path, laptop_business)
        assert result.valid_rows == 2
        assert result.total_rows == 2  # ردیف خالی شمرده نشد
    finally:
        Path(file_path).unlink(missing_ok=True)